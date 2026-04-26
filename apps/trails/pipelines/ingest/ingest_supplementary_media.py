import openpyxl
import sqlite3
import json
from pathlib import Path

DB_PATH = Path('retrieval/index/trails.db')
EXCEL_PATH = Path('corpus/raw/List of Trails Media.xlsx')

def ingest_excel_media():
    if not EXCEL_PATH.exists():
        print("Excel file not found.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Sheet names to process
    sheets = ["Main series games", "Spin-off games", "Anime", "Drama CDs"]
    
    media_count = 0
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Skipping sheet {sheet_name}: Not found.")
            continue
            
        sheet = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")
        
        # Iterate rows (skipping header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            title = row[0]
            if not title: continue
            
            # Simple ID generation
            mid = title.lower().replace(' ', '_').replace(':', '').replace('!', '').replace("'", "")
            mid = re.sub(r'[^a-zA-Z0-9_]', '', mid)[:50]
            
            # Determine type
            mtype = 'game' if "games" in sheet_name.lower() else sheet_name.lower().replace('s', '')
            
            # Get release date if available (usually col 2 or 3)
            # Depending on the sheet structure which I don't know yet, I'll be defensive
            release_date = str(row[1]) if len(row) > 1 else None
            
            cursor.execute('''
                INSERT OR IGNORE INTO media_registry (
                    media_id, media_type, english_title, release_date_en, is_main_series
                ) VALUES (?, ?, ?, ?, ?)
            ''', (mid, mtype, title, release_date, 1 if "main series" in sheet_name.lower() else 0))
            media_count += 1

    conn.commit()
    conn.close()
    print(f"[SUCCESS] Ingested {media_count} media entries from local Excel.")

if __name__ == "__main__":
    import re
    ingest_excel_media()
